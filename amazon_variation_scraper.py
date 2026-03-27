import os
import re
import sys
import json
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

def extract_all_amazon_cookies(headers):
    session = requests.Session()
    # 尝试处理可能出现的验证码页面
    response_initial = session.get('https://www.amazon.com/errors/validateCaptcha', headers=headers)

    soup = BeautifulSoup(response_initial.text, 'html.parser')

    amzn_input = soup.find('input', {'name': 'amzn'})
    amzn = amzn_input['value'] if amzn_input else None

    field_keywords_input = soup.find('input', {'name': 'field-keywords'})
    field_keywords = field_keywords_input['value'] if field_keywords_input else None

    if not amzn or not field_keywords:
        # Check if the response was the actual Amazon product page (no captcha)
        if 'captcha' not in response_initial.text.lower():
            # If no amzn/field-keywords found, but also no captcha, assume success and return session
            print("INFO: Initial request did not encounter a Captcha. Proceeding with the session.")
            return session.cookies.get_dict(), session
        
        print("错误: 未能在初始验证码页面找到 'amzn' 或 'field-keywords'。")
        return None, None # 返回 None, None 以便主程序处理

    params = {
        'amzn': amzn,
        'amzn-r': '/',
        'field-keywords': field_keywords  
    }

    # Submit the captcha form (even if field-keywords is empty/bogus)
    session.get('https://www.amazon.com/errors/validateCaptcha', params=params, headers=headers)

    final_cookies = session.cookies.get_dict()
    
    return final_cookies, session

def convert_json_to_excel(json_data_string, output_filename="Amazon_price_spider_US.xlsx", sheet_name="Sheet1"):
    """
    Converts a JSON string containing product dimension data into an Excel file
    using the openpyxl library.

    Args:
        json_data_string (str): A JSON string where keys are product ASINs
                                and values are lists containing dimension (e.g., "72\"Wx72\"L(Packof1)")
                                and color (e.g., "WhiteGreen").
        output_filename (str, optional): The name of the Excel file to create.
                                          Defaults to "Amazon_price_spider_US.xlsx".
        sheet_name (str, optional): The name of the sheet to write the data to.
                                    Defaults to "Sheet1".
    """
    try:
        # Parse the JSON string into a Python dictionary
        data = json.loads(json_data_string)

        # Check if the Excel file already exists
        if os.path.exists(output_filename):
            print(f"File '{output_filename}' exists. Loading workbook...")
            # Load the existing workbook
            wb = load_workbook(output_filename)
            
            # Check if the specified sheet exists
            if sheet_name in wb.sheetnames:
                print(f"Sheet '{sheet_name}' exists. Writing data...")
                ws = wb[sheet_name]
                # Clear existing content from the sheet before writing new data
                # Keep the header row if it exists, otherwise clear all
                if ws.max_row > 0:
                    ws.delete_rows(1, ws.max_row)
            else:
                print(f"Sheet '{sheet_name}' does not exist. Creating a new sheet...")
                # Create a new sheet with the specified name
                ws = wb.create_sheet(sheet_name)
        else:
            print(f"File '{output_filename}' does not exist. Creating a new workbook...")
            # Create a new workbook if the file doesn't exist
            wb = Workbook()
            # Remove the default sheet created by Workbook() if it's still there
            if 'Sheet' in wb.sheetnames:
                 std_ws = wb['Sheet']
                 wb.remove(std_ws)
                 
            ws = wb.create_sheet(sheet_name)


        # Define the header row
        headers = ['ASIN', 'color', 'N/A', 'Price', 'Coupon']
        ws.append(headers)

        # Iterate through the dictionary and write data to the worksheet
        for asin, details in data.items():
            size = ""
            color = ""
            
            # --- FIX: Handle different list lengths from dimensionValuesDisplayData ---
            if isinstance(details, list):
                if len(details) == 2:
                    # Assumes details[0] is Color, details[1] is Size (based on original comment logic)
                    color = details[0]
                    size = details[1]
                elif len(details) == 1:
                    # If only one attribute is present, place it in the 'Size' column for now
                    # This is a common case for simple variations (e.g., 'Pack of 4')
                    size = details[0] 
                else:
                    # Handle other cases or malformed data with more/less than 1 or 2 elements
                    print(f"Warning: Unexpected list length ({len(details)}) for ASIN {asin}: {details}. Appending only ASIN.")
            
            # Write data row, including empty strings for missing values
            ws.append([asin, size, color])

        # Save the workbook
        wb.save(output_filename)
        print(f"Data successfully saved to '{output_filename}' in sheet '{sheet_name}'.")

    except json.JSONDecodeError as e:
        print(f"Error decoding JSON: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def get_target_dict(html_content):
    # This regex is robust and correct for extracting the JSON
    pattern = r'"dimensionValuesDisplayData"\s*:\s*(\{.*?\})'

    # Search the text for the pattern.
    match = re.search(pattern, html_content, re.DOTALL)

    if match:
        # Extract the captured content (the JSON string).
        json_string = match.group(1)
        
        # Use the json module to parse the string into a Python dictionary.
        data = json.loads(json_string)
        return data
    else:
        print("Could not find dimensionValuesDisplayData in the text.")
        return {} # Return an empty dictionary on failure to prevent crashes

def main():
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8,zh-CN;q=0.7,zh;q=0.6',
        'device-memory': '8',
        'downlink': '1.45',
        'dpr': '1',
        'ect': '3g',
        'priority': 'u=0, i',
        'rtt': '300',
        'sec-ch-device-memory': '8',
        'sec-ch-dpr': '1',
        'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-ch-ua-platform-version': '"10.0.0"',
        'sec-ch-viewport-width': '2560',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
        'viewport-width': '2560',
    }
    cookies, session = extract_all_amazon_cookies(headers)
    if session is None:
        print("未能成功获取Amazon会话，程序退出。")
        sys.exit(1)

    time.sleep(2)
    asin = 'B0BLYSGFRW'
    sheet_name = asin

    print(f"\n--- Processing ASIN: {asin} in sheet: {sheet_name} ---")
    response = session.get(
        f'https://www.amazon.com/dp/{asin}',
        headers=headers
    )
    actual_product_data = get_target_dict(response.text)

    if not actual_product_data:
        print(f"Skipping ASIN {asin}: No dimension data found.")
        sys.exit(1)

    # Convert this inner dictionary to a JSON string
    json_string_for_conversion = json.dumps(actual_product_data)
    convert_json_to_excel(json_string_for_conversion,"Amazon_price_spider_US.xlsx", sheet_name)

if __name__ == '__main__':
    main()
