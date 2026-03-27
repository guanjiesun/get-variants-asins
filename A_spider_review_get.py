from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook # 导入 Workbook 类
import random
import re
def get_review_status(driver):
    """
    从页面提取评论数量并判断。
    大于 100 返回 1，小于等于 100 返回 0。
    """
    try:
        # 1. 设置显式等待，确保元素已加载（最多等待10秒）
        # 使用图片中显示的 data-hook 属性作为定位符，这是最准确的
        locator = (By.CSS_SELECTOR, 'div[data-hook="cr-filter-info-review-rating-count"]')
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(locator)
        )
        
        # 2. 获取文本内容 (例如: " 42 customer reviews ")
        raw_text = element.text
        
        # 3. 使用正则表达式提取数字（处理逗号分隔符，如 1,234）
        # \d+ 匹配数字，replace(',', '') 处理千分位
        match = re.search(r'([\d,]+)', raw_text)
        if match:
            count_str = match.group(1).replace(',', '')
            review_count = int(count_str)
            
            # 4. 逻辑判断
            return 1 if review_count > 100 else 0
        else:
            print("未能从文本中解析出数字")
            return 0
            
    except Exception as e:
        print(f"解析评论数量时出错: {e}")
        return 0

def random_sleep(min_time=1.5, max_time=2.5):
    """
    生成并执行一个随机浮点数的等待时间。
    这有助于模拟人类行为，减少被网站检测到的风险。
    默认等待时间在 0.8 秒到 2.5 秒之间。
    """
    sleep_time = random.uniform(min_time, max_time)
    print(f"-> 随机等待 {sleep_time:.2f} 秒...")
    time.sleep(sleep_time)
# ==============================================================================
# 核心排查与优化区域：连接到浏览器
# ==============================================================================
def attach_to_manual_browser(debugger_port=9222):
    """
    连接到通过远程调试端口启动的 Chrome 实例。
    """
    print(f"--- 尝试连接到端口 {debugger_port} 的已开启 Chrome 实例 ---")

    # TODO: 关键排查点 1 - 明确指定 ChromeDriver 路径
    # 请将此处的路径替换为您电脑上 chromedriver.exe 的实际存放路径！
    # 如果您已将 chromedriver 放在系统 PATH 中，可以保留 driver_path = 'chromedriver.exe'
    driver_path = 'c:/Program Files/Google/Chrome/Application/chromedriver.exe'  # 示例路径，请修改！
    
    try:
        # 1. 配置 Chrome 选项 (Options)
        chrome_options = Options()
        # 关键步骤：设置远程调试地址，指定要连接的浏览器实例
        chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{debugger_port}")

        # 2. 创建 Service 对象并实例化 WebDriver
        # 使用 Service 明确指定驱动路径，避免找不到驱动的问题
        service = Service(driver_path) 
        driver = webdriver.Chrome(service=service, options=chrome_options)
        
        print("\n*** 成功接管已开启的浏览器控制权！***")
        
        # 额外验证：获取当前打开的页面 URL
        current_url = driver.current_url
        print(f"当前浏览器 URL: {current_url}")
        
        return driver

    except Exception as e:
        print(f"\n连接失败。请检查以下几点：")
        print(f"1. **时序问题**：在运行此 Python 脚本前，**是否已启动 Chrome** 并使其窗口保持开启？")
        print(f"2. **端口冲突**：端口 {debugger_port} 是否被其他程序占用？")
        print(f"3. **驱动问题 (最常见)**：您是否在 driver_path='...' 处设置了正确的 ChromeDriver 路径？且驱动版本是否与您的 Chrome 浏览器版本兼容？")
        print(f"4. **启动参数**：您是否使用了完整的启动命令：{chrome_options.arguments[0]} --remote-debugging-port=9222 ...")
        print(f"错误信息: {e}")
        # 如果连接失败，返回 None 或直接退出，避免后续操作报错
        return None
    
# ==============================================================================
# 其它函数保持不变
# ==============================================================================

def click_choose_1(driver):
    wait = WebDriverWait(driver, 10)
    try:
        # Wait for the element to be clickable
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="reviewer-type-select"]//span[contains(@class, "cr-filter-dropdown")]/span/span')))
        button.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//ul/li/a[normalize-space(.)="Verified purchase only"]')))
        button.click()
        print('成功点击Verified purchase only')
        return 1,driver  # Return 1 if the button is found and clicked
    except :
        print('未找到按钮，无法点击')
        return 0,driver  # Return the driver if the button is not found
    
def click_choose_2(driver,i):
    wait = WebDriverWait(driver, 10)
    try:
        # Wait for the element to be clickable
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="star-rating-select"]//span[contains(@class, "cr-filter-dropdown")]/span/span')))
        button.click()
        TARGET_BUTTON_XPATH = '//ul/li[{0}][contains(@class, "star-filter-option")]/a[substring(normalize-space(.), string-length(normalize-space(.)) - string-length("star only") + 1) = "star only"][contains(@class, "a-dropdown-link")]'.format(i)
        button = wait.until(EC.element_to_be_clickable((By.XPATH, TARGET_BUTTON_XPATH)))
        button.click()                                          
        print('成功点击All star')
        return 1,driver  # Return 1 if the button is found and clicked
    except :
        print('未找到按钮，无法点击')
        return 0,driver  # Return the driver if the button is not found
    
def click_choose_3(driver):
    wait = WebDriverWait(driver, 10)
    try:
        # Wait for the element to be clickable
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="format-type-select"]//span[contains(@class, "cr-filter-dropdown")]/span/span')))
        button.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//ul/li/a[@id="format-type-dropdown_1"]')))
        button.click()
        print('成功点击All varient')
        return 1,driver  # Return 1 if the button is found and clicked
    except :
        print('未找到按钮，无法点击')
        return 0,driver  # Return the driver if the button is not found


def next_page(driver):
    wait = WebDriverWait(driver, 3)
    try:
        # Wait for the element to be clickable
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//li[@class="a-last"]')))
        button.click()
        print('成功点击next page')
        return 1,driver  # Return 1 if the button is found and clicked
    except :
        print('未找到next page按钮，无法点击')
        return 0,driver  # Return the driver if the button is not found

def extract_reviews_from_html(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    reviews = []

    # 查找所有评论块，根据resource.txt中的结构，每个评论都在一个 <li data-hook="review"> 标签中
    review_elements = soup.find_all('li', {'data-hook': 'review'})

    for review_element in review_elements:
        # 提取作者
        author_tag = review_element.find('span', class_='a-profile-name')
        author = author_tag.get_text(strip=True) if author_tag else 'N/A'

        # 提取评分
        rating_tag = review_element.find('i', class_='review-rating')
        rating_span = rating_tag.find('span', class_='a-icon-alt') if rating_tag else None
        rating = rating_span.get_text(strip=True) if rating_span else 'N/A'

        # 提取标题
        title_tag = review_element.find('a', {'data-hook': 'review-title'})
        title_span = title_tag.find('span', class_=None) if title_tag else None # 标题文本在a标签下的第二个span
        title = title_span.get_text(strip=True) if title_span else 'N/A'

        # 提取评论日期
        date_tag = review_element.find('span', {'data-hook': 'review-date'})
        review_date = date_tag.get_text(strip=True) if date_tag else 'N/A'

        # 提取颜色和尺寸
        format_strip_tag = review_element.find('a', {'data-hook': 'format-strip'})
        color = 'N/A'
        size = 'N/A'
        if format_strip_tag:
            format_text = format_strip_tag.get_text(separator='|', strip=True)
            parts = format_text.split('|')
            for part in parts:
                if 'Color:' in part:
                    color = part.replace('Color:', '').strip()
                elif 'Size:' in part:
                    size = part.replace('Size:', '').strip()

        # 提取评论正文
        body_tag = review_element.find('span', {'data-hook': 'review-body'})
        # 获取所有文本内容，并用换行符连接，同时处理 <br/> 标签
        review_body_parts = []
        if body_tag:
            for content in body_tag.contents:
                if content.name == 'br':
                    review_body_parts.append('\n')
                else:
                    review_body_parts.append(str(content).strip())
            review_body = ''.join(review_body_parts).strip()
            # 替换HTML实体，例如 &#034; 为 "，去除所有空格，并将换行符替换为 /
            review_body = BeautifulSoup(review_body, 'html.parser').get_text().replace('\n', '/')
        else:
            review_body = 'N/A'


        # 格式化输出字符串
        formatted_review = (
            f"{author}\n"
            f"{rating}\n"
            f"{title}\n"
            f"{review_date}\n"
            f"Color: {color}\n"
            f"Size: {size}\n"
            f"{review_body}\n"
        )
        reviews.append(formatted_review)

    return reviews

def transfer_data_range(target_file_path, target_sheet_name, extracted_reviews):
    try:
        # 尝试加载工作簿，如果文件不存在则创建一个新的
        try:
            target_wb = load_workbook(target_file_path)
        except FileNotFoundError:
            target_wb = Workbook()
            print(f"文件 '{target_file_path}' 不存在，已创建新文件。")

        # 获取或创建工作表
        if target_sheet_name in target_wb.sheetnames:
            target_sheet = target_wb[target_sheet_name]
            print(f"成功加载工作表: '{target_sheet_name}'")
        else:
            target_sheet = target_wb.create_sheet(target_sheet_name)
            print(f"工作表 '{target_sheet_name}' 不存在，已创建新工作表。")

        # 查找第一个空行作为起始行
        start_row = 1
        for i in range(1, target_sheet.max_row + 2): # 检查现有行和下一行
            target_cell = target_sheet.cell(row=i, column=1)
            if target_cell.value is None:
                start_row = i
                break
        
        # 如果是新工作表，写入表头
        if start_row == 1:
            headers = ["作者", "评分", "标题", "评论日期", "颜色", "尺寸", "评论正文"]
            for col_idx, header in enumerate(headers, 1):
                target_sheet.cell(row=start_row, column=col_idx, value=header)
            start_row += 1 # 表头占据一行，数据从下一行开始

        # 写入数据
        current_row = start_row
        for review_str in extracted_reviews:
            lines = review_str.strip().split('\n')
            
            author = lines[0] if len(lines) > 0 else 'N/A'
            rating = lines[1] if len(lines) > 1 else 'N/A'
            title = lines[2] if len(lines) > 2 else 'N/A'
            review_date = lines[3] if len(lines) > 3 else 'N/A'
            color = lines[4].replace('Color: ', '') if len(lines) > 4 and lines[4].startswith('Color:') else 'N/A'
            size = lines[5].replace('Size: ', '') if len(lines) > 5 and lines[5].startswith('Size:') else 'N/A'
            
            # 评论正文可能有多行
            review_body = '\n'.join(lines[6:]) if len(lines) > 6 else 'N/A'

            # 将数据写入行
            target_sheet.cell(row=current_row, column=1, value=author)
            target_sheet.cell(row=current_row, column=2, value=rating)
            target_sheet.cell(row=current_row, column=3, value=title)
            target_sheet.cell(row=current_row, column=4, value=review_date)
            target_sheet.cell(row=current_row, column=5, value=color)
            target_sheet.cell(row=current_row, column=6, value=size)
            target_sheet.cell(row=current_row, column=7, value=review_body)
            
            current_row += 1

        target_wb.save(target_file_path)
        print(f"数据传输完成。更改已保存到'{target_file_path}'。")

    except FileNotFoundError as e:
        print(f"错误: 文件未找到。请检查文件路径是否正确。详细信息: {e}")
    except KeyError as e:
        print(f"错误: 工作表名称 '{e}' 未找到。请检查工作表名称是否正确。")
    except Exception as e:
        print(f"发生了一个意外错误: {e}")


# ==============================================================================
# 主执行逻辑
# ==============================================================================

# 初始化 WebDriver
driver = attach_to_manual_browser()

# 检查是否成功连接到浏览器

if driver:
    ALL_ASIN = [
'B0C56Y2PS5',
'B0D24X1MVY',
'B0D25DJPFG',
'B0D24X1Y98',
'B0D25CX6QJ',
'B0DFHFRG73',
'B0D25HZY9Q',
'B0BLYX97B3',
'B0DFHJWBQ3',
'B0D9JHL5BL',
'B0CXCFD5HW',
'B0D5VRFXQK',
'B0DFHHLNN4',
'B0D9J4LS79',
'B0CXD345WP',
'B0BLYSMHQ3',
'B0D25LDRPS',
'B0C56WCVK6',
'B0C56YFDQN',
'B0C56ZNQVQ',
'B0DFHFL622',
'B0D5VQQBW1',
'B0C56Z36BJ',
'B0DFHFXSNS',
'B0BLYW1C38',
'B0D9JJSNC5',
'B0BLYSGFRW',
'B0D25GFQTD',
'B0C56YD1B6',
'B0D5VQPJRR',
'B0C56XBD8K',
'B0D25J19PM',
'B0DFHK39RF',
'B0BLYVMGYB',
'B0C56YLKZW',
'B0DFHFRRS6',
'B0D9J4LG7Y',
'B0DFHGR7YL',
'B0CXCR3MJZ',
'B0CXCX16Q5',
'B0CXCLXYG9',
'B0DFHG2YKM',
'B0D9JJ62NF',
'B0C56Z4KRF',
'B0D9JK62G9',
'B0C56Z3QLT',
'B0C56W9ZMM',
'B0D9JHBBCZ',
'B0CXCR5D36',
'B0D9JJZMML',
'B0C571R34Z',
'B0D5VP878W',
'B0DFHGZKGS',
'B0BLYW9CJZ',
]
    for ASIN in ALL_ASIN:
        driver.get(f"https://www.amazon.com/product-reviews/{ASIN}/ref=cm_cr_dp_d_show_all_btm?ie=UTF8")
        random_sleep()
        # check1, driver = click_choose_1(driver)
        # random_sleep()
        # if check1:
        check_0 = get_review_status(driver)
        if check_0:
            check2, driver = click_choose_3(driver)
            random_sleep() 
            check_0 = get_review_status(driver)
            if check2:
                if check_0:
                    for k in range(2, 7):
                        check3, driver = click_choose_2(driver,k)
                        random_sleep()
                        html_content = driver.page_source
                        reviews = extract_reviews_from_html(html_content)
                        transfer_data_range('reviews.xlsx', '评论数据', reviews)
                        for i in range(9):
                            check,driver = next_page(driver)
                            if check == 1:
                                random_sleep()
                                html_content = driver.page_source
                                reviews = extract_reviews_from_html(html_content)
                                transfer_data_range('reviews.xlsx', '评论数据', reviews)
                            else:
                                break
                else:
                    html_content = driver.page_source
                    reviews = extract_reviews_from_html(html_content)
                    transfer_data_range('reviews.xlsx', '评论数据', reviews)
                    for i in range(9):
                        check,driver = next_page(driver)
                        if check == 1:
                            random_sleep()
                            html_content = driver.page_source
                            reviews = extract_reviews_from_html(html_content)
                            transfer_data_range('reviews.xlsx', '评论数据', reviews)
                        else:
                            break
        else:
            html_content = driver.page_source
            reviews = extract_reviews_from_html(html_content)
            transfer_data_range('reviews.xlsx', '评论数据', reviews)
            for i in range(9):
                check,driver = next_page(driver)
                if check == 1:
                    random_sleep()
                    html_content = driver.page_source
                    reviews = extract_reviews_from_html(html_content)
                    transfer_data_range('reviews.xlsx', '评论数据', reviews)
                else:
                    break
else:
    print("未能连接到浏览器，脚本提前退出。")
   